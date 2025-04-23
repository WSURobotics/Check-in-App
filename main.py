import hidmsr.commands as cmds
import hidmsr.convert as conv
from datetime import datetime
from discord_webhook import DiscordWebhook
import os
import openpyxl

def extract_id(text: str) -> str:
    offset = 7
    index = text.find("6008770") + offset
    if index == offset-1:
        return ""
    return int(text[index:index + 8])

def main():
    data_dir = "Data"
    logs_dir = os.path.join(data_dir)
    os.makedirs(logs_dir, exist_ok=True)
    
    m = cmds.MSRDevice()
    m.set_hico()
    m.set_bpi(210, 75, 210)

    open_hours_path = os.path.join(logs_dir, "open_hours_log.xlsx")
    open_hours = openpyxl.load_workbook(open_hours_path)
    open_hours_sh = open_hours.active
    officer_list_path = os.path.join(logs_dir, "officer_list.xlsx")
    officer_list = openpyxl.load_workbook(officer_list_path)
    officer_list_sh = officer_list.active
    member_attendance_path = os.path.join(logs_dir, "member_attendance.xlsx")
    member_attendance = openpyxl.load_workbook(member_attendance_path)
    member_attendance_sh = member_attendance.active

    webhook = False
    webhook_url = "WEBHOOK URL HERE"
    webhook_in = DiscordWebhook(url=webhook_url, content="The club room is open")
    webhook_out = DiscordWebhook(url=webhook_url, content="The club room is closed")

    print(f"Ready to read cards. Writing to spreadsheets. Press Ctrl+C to exit.")

    try:
        while True:
            out = m.read()
            data = conv.decode_msr_data(out)
            card_id = extract_id(data)
            if card_id:
                current_date = datetime.now().strftime('%Y-%m-%d')
                current_time_str = datetime.now().strftime('%H:%M:%S')
                status = ""
                # Changing officer statuses
                officer_changed = False
                for row in officer_list_sh.iter_rows():
                    for cell in row:
                        if cell.value == card_id:
                            status_cell = officer_list_sh.cell(row=cell.row, column=cell.column + 1)
                            if status_cell.value == 'In':
                                status_cell.value = 'Out'
                                officer_changed = True
                                status = f"Officer {card_id} out"
                            else:
                                status_cell.value = 'In'
                                officer_changed = True
                                status = f"Officer {card_id} in"
                
                # Open hours logging using officer statuses
                if officer_changed == True:
                    officer_changed = False
                    all_out = True
                    for row in officer_list_sh.iter_rows():
                        for cell in row:
                            if cell.value == 'In':
                                all_out = False
                    last_status = open_hours_sh.cell(row=open_hours_sh.max_row, column=4).value
                    if last_status == 'In' and all_out == True:
                        # CODE FOR OUT
                        print('Room closed')
                        if webhook == True:
                            webhook_in.execute()
                        open_hours_sh.append([current_date, current_time_str, card_id, 'Out'])
                    elif last_status == 'Out' and all_out == False:
                        # CODE FOR IN
                        print('Room open')
                        if webhook == True:
                            webhook_out.execute()
                        open_hours_sh.append([current_date, current_time_str, card_id, 'In'])
                elif officer_changed == False:
                    # ADD NEW ROW WITH ID & TIME TO member_attendance_sh
                    member_attendance_sh.append([current_date, current_time_str, card_id])
                    status = f"Member attendance"
                print(f"Recorded ID: {card_id} at {current_date}, {current_time_str} - {status}")

                open_hours.save(open_hours_path)
                officer_list.save(officer_list_path)
                member_attendance.save(member_attendance_path)

    except KeyboardInterrupt:
            print("\nStopping card reader...")
    except Exception as e:
        print(f"\nError occurred: {str(e)}")
    finally:
        print("Program terminated.")
                    

if __name__ == "__main__":
    main()